import cv2
import numpy as np
import face_recognition
import os
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
import pandas as pd
from pathlib import Path



 
# path = 'image'
# images = []
# classNames = []
# myList = os.listdir(path)
# for cl in myList:
#     curImg = cv2.imread(f'{path}/{cl}')
#     images.append(curImg)
#     classNames.append(os.path.splitext(cl)[0])





# def timeInterval(s1,s2):

#     FMT = '%Y-%m-%d %H:%M:%S.%f'
    
#     tdelta = datetime.strptime(str(s1), FMT) - datetime.strptime(str(s2), FMT)
#     if tdelta.seconds>((3000/1000)%60):
#         print(tdelta)
#         return True
#     else: return False



# def createSheet():
#     path = ('name.xlsx')

#     wb_obj = openpyxl.load_workbook(path)
#     now = datetime.now()
#     name = now.strftime('%H:%M:%S')
#     wb_obj.create_sheet("New sheet")
#     wb_obj.save('name.xlsx') 
#     wb_obj.close()




# def createSheet():
#     path = ('name.xlsx')

#     wb_obj = openpyxl.load_workbook(path)
#     now = datetime.now()
#     #name = str(now.strftime('%dth%b, %Y'))
#     name = str(now.strftime('%dth%b, %Y'))

#     print(name)

#     ws = wb_obj.worksheets[-1]
#     sh = str(ws).split('"')
#     s = sh[1]
#     print(s+" Printing here")
#     if(name != s):
#         wb_obj.create_sheet(name)
#     else: name = s


#     print("done")

#     wb_obj.save('name.xlsx') 
    
#     ws = wb_obj.worksheets[-1]
#     sh = str(ws).split('"')
#     s = sh[1]
#     print(type(s))
#     wb_obj.close()
#     return s




# def test(name,result,sheetName):
#     path = ('name.xlsx')

#     wb_obj = openpyxl.load_workbook(path)
#     wb_obj.iso_dates = True
#     # highlight = NamedStyle(name="highlight", number_format='MM/DD/YYYY HH:MM:MM')
#     # wb_obj.add_named_style(highlight)
#     sheet = wb_obj[sheetName]


#     __row__ = sheet.max_row
#     column = sheet.max_column
#     print("column: " + str(column))
#     list = []
    
#     print("row num: "+ str(__row__))
#     for i in range(2,__row__+1):
#         list.append(sheet.cell(i,1).value)
    

#     row = __row__
#     now = datetime.now()
#     dtString = now.strftime("%H-%M-%S")
#     sheet.cell(1,1,value = "Name")
#     sheet.cell(1,1).font = Font(bold = True)

#     if name not in list:
#         sheet.cell(row+1,1,value=name)
#         sheet.cell(row+1,2,value = "Time")
#         sheet.cell(row+1,2).font = Font(bold = True)
#         sheet.cell(row+1,3,value=now) 
#         sheet.cell(row+1,3).number_format ="HH:MM:SS"

#         sheet.cell(row+2,2,value="Matching(%)")
#         sheet.cell(row+2,2).font = Font(bold = True)
#         sheet.cell(row+2,3,value=result)
          

#         print('test: '+ name + '\n' + 'Date: ' + dtString)
   
#     elif name in list:
#         timeList = []
#         index = list.index(name)
#         print("indext: "+str(index))
#         for i in range(1,column+1):
#             if ((sheet.cell(index+2,i).value != None)):
#                 timeList.append(sheet.cell(index+2,i).value)
        
#         print(timeList)
#         length = len(timeList)
#         print(timeList[length-1])

#         oldTime = timeList[length-1]
#         print("Time test: ")
#         print(oldTime)
#         print(type(oldTime))
#         print("Time test: ")
#         print(sheet.cell(index+2,column).value)
#         if timeInterval(now,oldTime) == True:
#             sheet.cell(index+2, length+1, value=now)
#             #sheet.cell(index+2, length+1).style = highlight
#             sheet.cell(index+2,length+1).number_format ="HH:MM:SS"
#             sheet.cell(index+3, length+1, value=result)  
            

         
#     wb_obj.save('name.xlsx') 
#     wb_obj.close()


def findEncodings(images): 
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList



def partho(): 
    path = 'photos'
    images = []
    classNames = []
    myList = os.listdir(path)
    for cl in myList:
        curImg = cv2.imread(f'{path}/{cl}')
        images.append(curImg)
        classNames.append(os.path.splitext(cl)[0])

    now = datetime.now()
    first = now.strftime("%H:%M:%S")
    print("encoding starting time " + first)

    encodeListKnown = findEncodings(images)
    now = datetime.now()
    second = now.strftime("%H-%M-%S")
    print('Encoding Complete ' + second)

def printH():
    print("Hello world")

# sheetName = createSheet()
# print("test01")
# cap = cv2.VideoCapture(0)

# while True:
#     success, img = cap.read()
#     #img = captureScreen()
#     imgS = cv2.resize(img,(0,0),None,0.25,0.25)
#     imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
    
#     facesCurFrame = face_recognition.face_locations(imgS)
#     encodesCurFrame = face_recognition.face_encodings(imgS,facesCurFrame)
#     for encodeFace,faceLoc in zip(encodesCurFrame,facesCurFrame):
#         matches = face_recognition.compare_faces(encodeListKnown,encodeFace)
#         faceDis = face_recognition.face_distance(encodeListKnown,encodeFace)
#         matchIndex = np.argmin(faceDis)

#         print(faceDis[matchIndex])
#         if(faceDis[matchIndex]<0.50):
#             result = faceDis[matchIndex]*100
#             result = 100 - float(result)

#             if matches[matchIndex]:
#                 name = classNames[matchIndex].upper()
#                 show = 'Matching = ' + str(int(result)) + '%  ' +name 
#                 y1,x2,y2,x1 = faceLoc
#                 y1, x2, y2, x1 = y1*4,x2*4,y2*4,x1*4
#                 cv2.rectangle(img,(x1,y1),(x2,y2),(0,255,0),2)
#                 cv2.rectangle(img,(x1,y2-35),(x2,y2),(0,255,0),cv2.FILLED)
#                 cv2.putText(img, show,(x1+6,y2-6),cv2.FONT_HERSHEY_COMPLEX,0.5,(255,255,255),2)

#                 test(name,result,sheetName)
#                 # markAttendance(name,result)
                
                
#     cv2.imshow('Webcam',img)
#     if cv2.waitKey(2) == 27:
#         break


# # import cv2
# # import numpy as np
# # import face_recognition
# # import os
# # import time
# # import csv
# # from datetime import datetime
# # import openpyxl
# # from pathlib import Path


# # c = 0 
# # path = 'img'
# # images = []
# # classNames = []
# # myList = os.listdir(path)
# # for cl in myList:
# #     curImg = cv2.imread(f'{path}/{cl}')
# #     images.append(curImg)
# #     classNames.append(os.path.splitext(cl)[0])

# # now = datetime.now()
# # dtString = now.strftime('%H:%M:%S')
# # print("encoding starting time " + dtString)

# def findEncodings(images):
#     encodeList = []
#     for img in images:
#         img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
#         encode = face_recognition.face_encodings(img)[0]
#         encodeList.append(encode)
#     return encodeList

# def timeInterval(s1,s2):

#     FMT = '%H:%M:%S'
#     tdelta = datetime.strptime(s2, FMT) - datetime.strptime(s1, FMT)
#     if tdelta>2:
#         return True
#     else: return False    

# nameList = []
# time = []

# def excelFile(name,result):
#     path = ('home.xlsx')

#     wb_obj = openpyxl.load_workbook(path)
#     sheet = wb_obj.active
#     print(sheet.max_row, sheet.max_column)

#     list = []
#     for row in sheet.iter_rows(max_row=6):
#         for cell in row:
#             list.append(cell.value) 
#             # print(cell.value)
#         print(list)




# def markAttendance(name,result):
#     global c
#     with open('Book1.csv','r+',encoding='utf-8-sig',newline='') as f:
#         myDataList = f.readlines()
        
        

#         for line in myDataList:
#             # print(line)
#             entry = line.split(',')
#             # if entry.startswith('\ufeff'):
#             #     entry = entry.encode('utf8')[3:].decode('utf8')

#             # print(entry[0])

#             nameList.append(entry[0])
#             nameList.append(entry[2])
            
#             # #time.append(entry[2])
#             # c+=1
#             # print(str(len(nameList))+" ")
#             # print(nameList)
        
#         now = datetime.now()
#         dtString = now.strftime('%H:%M:%S') 
#         result = str(result)+'\n' 
#         if ("      "+name) not in nameList:

#             #f.write(f' {name},{dtString},{result} ')
            
#             #f.writelines(f'{name},{dtString}')
#             print(name)
#             #print('Name: '+ name + '\n' + 'Date: ' + dtString)
#         elif  ("      "+name) in nameList:
#             if  timeInterval(time in entry[2] ,now) == True:
#             #     f.write(f' {name},{dtString},{result} ')
#                 print('Name: '+ name + '\n' + 'Date: ' + dtString)
                
#                 print(entry[2])
#         # count = nameList[0].index("\ufeff "+name)
#         # print(count)
#         # print(nameList[count][2])
#         nameList.clear()
#         f.close() 
        
             
           
 
# encodeListKnown = findEncodings(images)
# now = datetime.now()
# dtString = now.strftime('%H:%M:%S')
# print('Encoding Complete' + dtString)
 
# cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
 
# while True:
#     success, img = cap.read()
#     #img = captureScreen()
#     imgS = cv2.resize(img,(0,0),None,0.25,0.25)
#     imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
    
#     facesCurFrame = face_recognition.face_locations(imgS)
#     encodesCurFrame = face_recognition.face_encodings(imgS,facesCurFrame)
#     for encodeFace,faceLoc in zip(encodesCurFrame,facesCurFrame):
#         matches = face_recognition.compare_faces(encodeListKnown,encodeFace)
#         faceDis = face_recognition.face_distance(encodeListKnown,encodeFace)
#         matchIndex = np.argmin(faceDis)

#         print(faceDis[matchIndex])
#         if(faceDis[matchIndex]<0.50):
#             result = faceDis[matchIndex]*100
#             result = 100 - float(result)

#             if matches[matchIndex]:
#                 name = classNames[matchIndex].upper()
#                 show = 'Matching = ' + str(int(result)) + '%  ' +name 
#                 y1,x2,y2,x1 = faceLoc
#                 y1, x2, y2, x1 = y1*4,x2*4,y2*4,x1*4
#                 cv2.rectangle(img,(x1,y1),(x2,y2),(0,255,0),2)
#                 cv2.rectangle(img,(x1,y2-35),(x2,y2),(0,255,0),cv2.FILLED)
#                 cv2.putText(img, show,(x1+6,y2-6),cv2.FONT_HERSHEY_COMPLEX,0.5,(255,255,255),2)

                
#                 #time.sleep(2)
#                 markAttendance(name,result)
                
#                 # testing(name,result)
#     cv2.imshow('Webcam',img)
#     if cv2.waitKey(2) == 27:
#         break



# import openpyxl
# import pandas as pd
# from pathlib import Path

# path = ('home.xlsx')

# wb_obj = openpyxl.load_workbook(path)
# sheet = wb_obj.active
# # print(sheet.max_row, sheet.max_column)
# row = sheet.max_row
# col = sheet.max_column
# list = []
# count = 0
# print(row)
# for i in range(2,row+1):
#     # for j in range(1,col+1):
#     #     str = sheet.cell(i,j).value
#     #     print(count)
#     #     print(str)
#     #     count+=1
#     list.append(sheet.cell(i,1).value)
#     list.append(sheet.cell(i,2).value)
#     list.append(sheet.cell(i,3).value)

# for i in range(2,row+1):
#     # for j in range(1,col+1):
#     #     str = sheet.cell(i,j).value
#     #     print(count)
#     #     print(str)
#     #     count+=1

#     list.append(sheet.cell(i,1).value)
#     list.append(sheet.cell(i,2).value)
#     list.append(sheet.cell(i,3).value)




# df = pd.read_excel (r'home.xlsx') #place "r" before the path string to address special character, such as '\'. Don't forget to put the file name at the end of the path + '.xlsx'
# print (df)


