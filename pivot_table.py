import pandas as pd
import cv2
import numpy as np
import face_recognition
import os
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from pathlib import Path


# Location of the file
# loc = "name.xlsx"

# # Reading the file
# df=pd.read_excel(loc,sheet_name='19thAug, 2021')
# # # s = df.iat[2,2].strftime("%H:%M:%S")
# # # print(s)
# # # #Transposing the file
# # # # out = df.pivot_table(values='H', index=['V','ED','D','F','SF','T'],columns=['SPH'], fill_value=0)
# # # # print(out)

# out1 = pd.DataFrame(df)

# out2 = pd.DataFrame(out1.transpose())


# # # # out2.rename(columns={'SPH':-1}, inplace=True)
# # # # print(out2)

# # # # out2.rename(columns=lambda x: x+1, inplace=True)
# # # # print(out2)

# # # # out2.rename(columns={0:'Name'}, inplace=True)
# # # # print(out2)

# #Save
# out2.to_excel("analysis.xlsx",sheet_name = 'Sheet1',index=False)

wb_obj = openpyxl.load_workbook("analysis.xlsx")
wb_obj.iso_dates = True

sheet = wb_obj['Sheet1']

row = sheet.max_row
column = sheet.max_column
for i in range(4,row+1):
    for j in range(1,column+1,2):
    
        # s = df.iat[2,2].strftime("%H:%M:%S")
        # sheet.cell(i,j).number_format ="HH:MM:SS"
        # s = sheet.cell(i,j).value
        s = (str(sheet.cell(i,j).value).split(" ")[-1]).split(".")[0]
        sheet.cell(i,j,value=s)
        # s = s.split(" ")
        print(s,end=" ")
    print()
# s = (str(sheet.cell(5,1).value).split(" ")[-1]).split(".")[0]

wb_obj.save('analysis.xlsx')
wb_obj.close()
print("done")
